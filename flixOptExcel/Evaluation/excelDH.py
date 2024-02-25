import pandas as pd
import numpy as np
import os.path
from openpyxl import load_workbook
from pathlib import Path

from flixOptExcel.Evaluation.flixPostprocessingXL import flixPostXL
from flixOptExcel.Evaluation.HelperFcts_post import resample_data, rs_in_two_steps, reorder_columns


def run_excel_graphics_main(calc: flixPostXL, custom_output_file_path: str = "default"):
    """
    Generate annual comparison plots and save them to an Excel workbook.

    Parameters:
    - calc (flixPostXL): Solved calculation of type flixPostXL.
    - custom_output_file_path (str): A custom output file path (full path) for the Excel workbook. Default: Same as calc.

    Returns:
    None

    The function generates various annual comparison plots based on the provided calculation results (flixPostXL)
    and saves them to an Excel workbook. The workbook is created using a template specified in the calculation.

    Plots and corresponding data are organized into different sheets within the workbook:
    - "Waermelast und Verluste": Fernwärme load and losses data.
    - "Kostenübersicht": Costs overview data.
    - "Wärmeerzeugung": Fernwärme generation data.
    - "Installierte Leistung": Installed capacity data.
    - "Wärmevollkosten": Heat full costs data.
    - "Wärmekosten Variabel": Variable heat costs data.
    - "Emissionen": Emissions data.
    - "Energieträger": Energy carriers data.
    - "Stromerzeugung": Electricity generation data.
    - "Speicherkapazität": Storage capacity data.
    - "Speicher Summen": Summed storage fill level data.

    The Excel workbook is saved in the same folder as the calculation results with the filename
    "Jahresübersicht-{calc.infos['calculation']['name']}.xlsx". If a custom_output_file_path is provided,
    the workbook is saved at that location.

    Example:
    ```
    calc = flixPostXL(...)  # Create or obtain flixPostXL instance
    run_excel_graphics_main(calc)  # Save the workbook in the default location
    run_excel_graphics_main(calc, custom_output_file_path="path/to/save/file.xlsx")  # Save the workbook in a custom location
    ```

    """
    print("Overview Plots to Excel...")

    if custom_output_file_path == "default":
        output_file_path = calc.folder
    else:
        output_file_path = custom_output_file_path

    templ_path_excel_main = Path(__file__).resolve().parent.parent / "resources" / "Template_Evaluation_Overview.xlsx"

    wb = load_workbook(templ_path_excel_main)
    filename = f"Jahresübersicht-{calc.infos['calculation']['name']}.xlsx"
    path_excel_main = os.path.join(output_file_path, filename)
    wb.save(path_excel_main)

    excel = cExcelFcts(calc)

    with pd.ExcelWriter(path_excel_main, mode="a", engine="openpyxl", if_sheet_exists='overlay') as writer:
        df = excel.get_fernwaerme_last_and_loss("YE", "sum")
        df.to_excel(writer, index=True, sheet_name="Waermelast und Verluste")

        df = excel.get_costs_and_funding("YE")
        df.to_excel(writer, index=True, sheet_name="Kostenübersicht")

        df = excel.get_fernwaerme_erz("YE", "mean")
        df.to_excel(writer, index=True, sheet_name="Wärmeerzeugung")

        df = excel.get_installierte_leistung(resamply_by="YE", rs_method="mean", flows=True, storage_capacity=False,
                                             grouped=True, actual_storage_capacity=False)
        df.to_excel(writer, index=True, sheet_name="Installierte Leistung")

        df = excel.get_waermekosten(with_fix_costs=True, resamply_by="YE")
        df.to_excel(writer, index=True, sheet_name="Wärmevollkosten")

        df = excel.get_waermekosten(with_fix_costs=False, resamply_by="YE")
        df.to_excel(writer, index=True, sheet_name="Wärmekosten Variabel")

        df = excel.get_emissions(resamply_by="YE", rs_method="sum")
        df.to_excel(writer, index=True, sheet_name="Emissionen")

        df = excel.get_eingesetzte_energietraeger(resamply_by="YE", rs_method="mean")
        df.to_excel(writer, index=True, sheet_name="Energieträger")

        df = excel.get_stromerzeugung(resamply_by="YE")
        df.to_excel(writer, index=True, sheet_name="Stromerzeugung")

        df_speicher_kapazitaet_Y = excel.get_speicher_kapazitaet(resamply_by="YE",
                                                                 grouped=True, actual_storage_capacity=False)
        df_speicher_kapazitaet_Y.to_excel(writer, index=True, sheet_name="Speicherkapazität")

        df_speicher_fuellstand_sum_H = excel.get_speicher_fuellstand("h", "mean", allocated=True).reset_index(drop=True)
        df_speicher_fuellstand_sum_H.to_excel(writer, index=True, sheet_name="Speicher Summen")

    print("...Overview Plots to Excel finished")


def run_excel_graphics_years(calc: flixPostXL, short_version = False, custom_output_file_path: str = "default"):
    """
    Generate detailed annual comparison plots and save them to individual Excel workbooks for each year.

    Parameters:
    - calc (flixPostXL): Solved calculation of type flixPostXL.
    - short_version (bool): If True, generate a shortened version of the plots. Default is False.
    - custom_output_file_path (str): A custom output folder path for the Excel workbooks. default: Same as calc.

    Returns:
    None

    The function generates detailed annual comparison plots based on the provided calculation results (flixPostXL)
    and saves them to individual Excel workbooks. The workbooks are created using a template specified in the calculation.

    Plots and corresponding data are organized into different sheets within each workbook for the specified years:
    - "Wärmeerzeugung": Fernwärme generation data.
    - "Installierte Leistung": Installed capacity data.
    - "Wärmevollkosten": Heat full costs data.
    - "Wärmekosten Variabel": Variable heat costs data.
    - "Emissionen": Emissions data.
    - "Energieträger": Energy carriers data.
    - "Stromerzeugung": Electricity generation data.
    - "Speicherkapazität": Storage capacity data.
    - "Speicherfüllstand D": Daily storage fill level data.
    - "SpeicherFlows D": Daily storage flows data.
    - "WärmeErz-Last-D": Duration curve for heat generation and electricity prices, sorted by heat demand (Daily mean values).
    - "WärmeErz-Strom-D": Duration curves for heat generation and electricity prices, sorted by electricity prices (Daily mean values).

    If short_version is False (default), additional sheets are generated:
    - "WärmeErz-Last": Duration curve for heat generation and electricity prices, sorted by heat demand (Hourly values).
    - "WärmeErz-Strom": Duration curves for heat generation and electricity prices, sorted by electricity prices (Hourly values).
    - "Wärmeerzeugung_Februar": Hourly data for heat generation in February.
    - "Wärmeerzeugung_Juli": Hourly data for heat generation in July.
    - "WärmeErz-Last-DL-H": Annual load duration curves for heat generation. (Individually sorted for every generator)
    - "Speicher Summen": Hourly storage fill level data (Allocated over all storages).
    - "Speicherfüllstand H": Hourly storage fill level data for each individual storage.

    The Excel workbooks are saved in the specified output folder with filenames like
    "Jahr_{year}-{calc.infos['calculation']['name']}.xlsx".

    Example:
    ```
    calc = flixPostXL(...)  # Create or obtain flixPostXL instance
    run_excel_graphics_years(calc)  # Save the detailed workbooks in the default location
    run_excel_graphics_years(calc, short_version=True)  # Save shortened version of the workbooks in the default location
    run_excel_graphics_years(calc, custom_output_file_path="path/to/save/folder")  # Save the detailed workbooks in a custom location
    ```

    """
    if custom_output_file_path == "default":
        output_file_path = calc.folder
    else:
        output_file_path = custom_output_file_path

    print("Annual Plots to Excel...")
    excel = cExcelFcts(calc)

    # computation for the whole calculation

    df_fernwaerme_erz_nach_techn_D = excel.get_fernwaerme_erz(resamply_by="d", rs_method="mean")  # Wärmeerzeugung

    df_installierte_leistung_Y = excel.get_installierte_leistung(resamply_by="YE", rs_method="mean", flows=True,
                                                                 storage_capacity=False,
                                                                 grouped=True, actual_storage_capacity=False)

    df_waermekosten_vollkosten_D = excel.get_waermekosten(with_fix_costs=True, resamply_by="d")

    df_waermekosten_varCosts_D = excel.get_waermekosten(with_fix_costs=False, resamply_by="d")

    df_emissions_D = excel.get_emissions(resamply_by="d", rs_method="sum")

    df_eingesetzte_energietraeger_D = excel.get_eingesetzte_energietraeger(resamply_by="d", rs_method="mean")

    df_stromerzeugung_D = excel.get_stromerzeugung(resamply_by="d")

    df_speicher_kapazitaet_D = excel.get_speicher_kapazitaet(resamply_by="d",
                                                             grouped=True, actual_storage_capacity=True)

    df_speicher_fuellstand_D = excel.get_speicher_fuellstand("d", "mean", allocated=False)

    df_speicher_flows_D = excel.get_speicher_flows("d", "mean", allocated=False)

    print("......computation of data for short version finished")
    if not short_version:
        df_fernwaerme_erz_nach_techn_H = excel.get_fernwaerme_erz(resamply_by="h", rs_method="mean")
        df_speicher_fuellstand_H = excel.get_speicher_fuellstand("h", "mean", allocated=False)
        df_speicher_fuellstand_H_alloc = excel.get_speicher_fuellstand("h", "mean", allocated=True)

    # TODO: weitere Grafiken

    print("......computation of data finished")

    templ_path_excel_year = Path(__file__).resolve().parent.parent / "resources" / "Template_Evaluation_Year.xlsx"

    for index, year in enumerate(excel.calc.years):
        wb = load_workbook(templ_path_excel_year)
        filename = f"Jahr_{year}-{excel.calc.infos['calculation']['name']}.xlsx"
        path_excel_year = os.path.join(output_file_path, filename)
        wb.save(path_excel_year)

        with pd.ExcelWriter(path_excel_year, mode="a", engine="openpyxl", if_sheet_exists='overlay') as writer:
            # Wärmeerzeugung nach Technologie
            df = df_fernwaerme_erz_nach_techn_D[df_fernwaerme_erz_nach_techn_D.index.year == year]
            df.to_excel(writer, index=True, sheet_name="Wärmeerzeugung")

            # Installierte Leistung nach Technologie
            df = df_installierte_leistung_Y[df_installierte_leistung_Y.index == year]
            df.to_excel(writer, index=True, sheet_name="Installierte Leistung")

            # Wärmevollkosten
            df = df_waermekosten_vollkosten_D[df_waermekosten_vollkosten_D.index.year == year]
            df.to_excel(writer, index=True, sheet_name="Wärmevollkosten")

            # Wärmekosten Betrieb
            df = df_waermekosten_varCosts_D[df_waermekosten_varCosts_D.index.year == year]
            df.to_excel(writer, index=True, sheet_name="Wärmekosten Variabel")

            # Emissionen
            df = df_emissions_D[df_emissions_D.index.year == year]
            df.to_excel(writer, index=True, sheet_name="Emissionen")

            # Energieträger
            df = df_eingesetzte_energietraeger_D[df_eingesetzte_energietraeger_D.index.year == year]
            df.to_excel(writer, index=True, sheet_name="Energieträger")

            # Stromerzeugung
            df = df_stromerzeugung_D[df_stromerzeugung_D.index.year == year]
            df.to_excel(writer, index=True, sheet_name="Stromerzeugung")

            # Speicherkapazität allokiert
            df = df_speicher_kapazitaet_D[df_speicher_kapazitaet_D.index.year == year]
            df.to_excel(writer, index=True, sheet_name="Speicherkapazität")

            # Speicherfüllstand nicht allokiert (Tageswerte)
            df = df_speicher_fuellstand_D[df_speicher_fuellstand_D.index.year == year]
            df.to_excel(writer, index=True, sheet_name="Speicherfüllstand D")

            # Speicherflows nicht allokiert (Tageswerte)
            df = df_speicher_flows_D[df_speicher_flows_D.index.year == year]
            df.to_excel(writer, index=True, sheet_name="SpeicherFlows D")

            # Wärmeerzeugung als Jahresdauerlinien (Tagesmittelwerte)
            df = df_fernwaerme_erz_nach_techn_D[df_fernwaerme_erz_nach_techn_D.index.year == year]
            df.sort_values("Wärmelast", ascending=False,ignore_index=True).to_excel(writer, index=True, sheet_name="WärmeErz-Last-D")
            df.sort_values("Strompreis", ascending=False,ignore_index=True).to_excel(writer, index=True, sheet_name="WärmeErz-Strom-D")

            print(f"......Year-{year} finished (short version)")
            if not short_version:
                # Wärmeerzeugung als Jahresdauerlinien (Stundenwerte)
                df = df_fernwaerme_erz_nach_techn_H[df_fernwaerme_erz_nach_techn_H.index.year == year]
                df.sort_values("Wärmelast", ascending=False, ignore_index=True).to_excel(writer, index=True, sheet_name="WärmeErz-Last")
                df.sort_values("Strompreis", ascending=False, ignore_index=True).to_excel(writer, index=True, sheet_name="WärmeErz-Strom")

                # Wärmeerzeugung im Februar und Juli (Stundenwerte)
                df = df_fernwaerme_erz_nach_techn_H[df_fernwaerme_erz_nach_techn_H.index.year == year]
                df.loc[df.index.month == 2].to_excel(writer, index=True, sheet_name="Wärmeerzeugung_Februar")
                df.loc[df.index.month == 7].to_excel(writer, index=True, sheet_name="Wärmeerzeugung_Juli")

                # Jahresdauerlinien der einzelnen Wärmeerzeuger (Stundenwerte)
                df = df_fernwaerme_erz_nach_techn_H[df_fernwaerme_erz_nach_techn_H.index.year == year]
                df = pd.DataFrame(-np.sort(-df.values, axis=0), columns=df.columns)

                df.to_excel(writer, index=True, sheet_name="WärmeErz-Last-DL-H")

                # Speicherfüllstand (Stundenwerte) allokiert
                df = df_speicher_fuellstand_H_alloc[df_speicher_fuellstand_H_alloc.index.year == year]
                df.to_excel(writer, index=True, sheet_name="Speicher Summen")

                # Speicherfüllstand (Stundenwerte) nicht allokiert
                df = df_speicher_fuellstand_H[df_speicher_fuellstand_H.index.year == year]
                df.to_excel(writer, index=True, sheet_name="Speicherfüllstand H")
            print(f"...Year-{year} finished")

            # TODO: weitere Grafiken

    print("...Annual Plots to Excel finished")


class cExcelFcts():
    def __init__(self, calc: flixPostXL):
        self.calc = calc

    def get_costs_and_funding(self, resamply_by):
        funding_var = self.calc.get_effect_results("funding", origin="operation", as_TS=True)
        funding_fix = self.calc.get_effect_results("funding", origin="invest", as_TS=True)
        costs_var = self.calc.get_effect_results("costs", origin="operation", as_TS=True)
        costs_fix = self.calc.get_effect_results("costs", origin="invest", as_TS=True)

        df = pd.DataFrame(data={"Fixkosten": costs_fix,
                                "Variable Kosten": costs_var,
                                "Förderung Invest": -funding_fix,
                                "Förderung Betrieb": -funding_var},
                          index=self.calc.timeSeries,
                          )
        df = resample_data(df, self.calc.years, resamply_by, "sum")
        return df

    def get_fernwaerme_erz(self, resamply_by, rs_method):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling

            if "d", Strompreis and Wärmelast are added to the DataFrame in first and second column
            if "YE", Wärmelast and sorages are not included
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value

        Returns
        -------
        pd.DataFrame
        '''

        if resamply_by == "YE":
            df_fernwaerme = self.calc.to_dataFrame("Fernwaerme", "inout", grouped=False)  # ohne Wärmelast, ohne Speicher
            df_fernwaerme_grouped = self.calc.group_df_by_mapping(df_fernwaerme)
            df_fernwaerme_grouped.drop(columns=["Wärmelast_mit_Verlust"], inplace=True)
        else:
            df_fernwaerme_grouped = self.calc.to_dataFrame("Fernwaerme", "inout", grouped=True)
            df_fernwaerme_grouped["Wärmelast_mit_Verlust"] = -1 * df_fernwaerme_grouped["Wärmelast_mit_Verlust"]  # reinverting
            df_fernwaerme_grouped = pd.concat([df_fernwaerme_grouped, self.calc.getFuelCosts()["Strompreis"]], axis=1)

        df_fernwaerme_erz_nach_techn = resample_data(df_fernwaerme_grouped, self.calc.years, resamply_by, rs_method)

        df_fernwaerme_erz_nach_techn = self.merge_into_dispatch_structure(df_fernwaerme_erz_nach_techn)

        return df_fernwaerme_erz_nach_techn

    def get_installierte_leistung(self, resamply_by, rs_method, flows: bool, storage_capacity: bool, grouped=False,
                                  actual_storage_capacity: bool = False):
        '''
        Parameters
        ----------
        calc1 : flix_results
            Calculation Object
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling

        Returns
        -------
        pd.DataFrame
        '''
        df_invest = self.calc.get_invest_results_as_TS(flows=flows, storages=storage_capacity,
                                                       grouped=grouped, actual_storage_capacity=actual_storage_capacity)
        df_invest = reorder_columns(df_invest)

        if df_invest.empty:
            return df_invest
        else:
            df_invest = resample_data(df_invest, self.calc.years, resamply_by, rs_method)
            df_invest = self.merge_into_dispatch_structure(df_invest)
            return df_invest

    def get_waermekosten(self, with_fix_costs, resamply_by):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value

        Returns
        -------
        pd.DataFrame
        '''
        heat = self.calc.to_dataFrame("Waermelast", "in")

        if with_fix_costs:
            costs_total = pd.Series(self.calc.get_effect_results(effect_name="costs", origin="all", as_TS=True),
                                    index=self.calc.timeSeries)
        else:
            costs_total = pd.Series(self.calc.get_effect_results(effect_name="costs", origin="operation", as_TS=True),
                                    index=self.calc.timeSeries)

        # Unterschiedung zwischen Resampling
        if resamply_by == "d":
            rs_method_base = "h"
            new_columns = ["Tagesmittel", "Minimum (Stunde)", "Maximum (Stunde)"]
        elif resamply_by == "YE":
            rs_method_base = "d"
            new_columns = ["Jahresmittel", "Minimum (Tagesmittel)", "Maximum (Tagesmittel)"]
        else:
            raise ValueError(f"not implemented for resamply_by parameter: '{resamply_by}'")

        mean_costs_increment = resample_data(costs_total, self.calc.years, rs_method_base, "mean").iloc[:, 0]
        mean_heat_increment = resample_data(heat, self.calc.years, rs_method_base, "mean").iloc[:, 0]
        mean_costs_per_heat_increment = pd.DataFrame(mean_costs_increment / mean_heat_increment,
                                                     columns=["EURvarPerMWh"])

        minY = resample_data(mean_costs_per_heat_increment, self.calc.years, resamply_by, "min", rs_method_base)
        maxY = resample_data(mean_costs_per_heat_increment, self.calc.years, resamply_by, "max", rs_method_base)
        increment_sum_of_costs_total = resample_data(mean_costs_increment, self.calc.years, resamply_by, "sum",
                                                     rs_method_base).iloc[:, 0]
        increment_sum_of_heat_total = resample_data(mean_heat_increment, self.calc.years, resamply_by, "sum", rs_method_base).iloc[:,
                                      0]
        meanY = increment_sum_of_costs_total / increment_sum_of_heat_total

        df = pd.concat([meanY, minY, maxY], axis=1)
        df.columns = new_columns

        return df

    def get_emissions(self, resamply_by, rs_method):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value

        Returns
        -------
        pd.DataFrame
        '''
        heat = self.calc.to_dataFrame("Waermelast", "in")

        CO2 = pd.DataFrame(self.calc.get_effect_results(effect_name="CO2FW", origin="operation", as_TS=True),
                           index=self.calc.timeSeries)

        CO2_per_increment = resample_data(CO2, self.calc.years, resamply_by, rs_method).iloc[:, 0]
        heat_per_increment = resample_data(heat, self.calc.years, resamply_by, rs_method).iloc[:, 0]
        CO2_per_heat = CO2_per_increment / heat_per_increment * 1000  # from t/MWh to kg/MWh
        df_emissions = pd.concat([CO2_per_heat.round(1), CO2_per_increment, heat_per_increment], axis=1)
        df_emissions.columns = ["kgCO2PerMWh", "tCO2absolut", "MWhabsolut"]

        return df_emissions

    def get_eingesetzte_energietraeger(self, resamply_by, rs_method):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value

        Returns
        -------
        pd.DataFrame
        '''
        df_sources = self.calc.get_sources_and_sinks(sources=True, sinks=False, sinks_n_sources=False)
        df = resample_data(df_sources, self.calc.years, resamply_by, rs_method)
        df = reorder_columns(df)

        return df

    def get_stromerzeugung(self, resamply_by):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value

        Returns
        -------
        resampled DataFrame with new columns:
            if resamply_by = "d": ["Tagesmittel", "Minimum (Stunde)", "Maximum (Stunde)"]
            if resamply_by = "YE": ["Jahresmittel", "Minimum (Tagesmittel)", "Maximum (Tagesmittel)"],
        '''
        df_stromerzeugung = self.calc.to_dataFrame("StromEinspeisung", "out",invert_Output=False)
        df = rs_in_two_steps(df_stromerzeugung, self.calc.years, resamply_by, "h")

        return df

    def get_speicher_kapazitaet(self, resamply_by, grouped, actual_storage_capacity: bool):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling

        Returns
        -------
        resampled DataFrame with capacity of all Storages
        '''
        invest_results_speicher = self.calc.get_invest_results_as_TS(storages=True, flows=False,
                                                                     actual_storage_capacity=actual_storage_capacity)
        if invest_results_speicher.empty:
            invest_results_speicher = pd.DataFrame(np.zeros(len(self.calc.timeSeries)), index=self.calc.timeSeries)
            invest_results_speicher.rename(columns={invest_results_speicher.columns[0]: "Speicher"}, inplace=True)

        elif grouped:
            invest_results_speicher = self.calc.group_df_by_mapping(invest_results_speicher)

        df = resample_data(invest_results_speicher, self.calc.years, resamply_by, "max")

        return df

    def get_speicher_fuellstand(self, resamply_by, rs_method, allocated):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value
        allocated : boolean
            True: sum of all storages in column one and netto flow of all Storages in column 2
            False: charge state of storage separately, no flow values

        Returns
        -------
        resampled DataFrame with total charge_state of all Storages
        '''
        df_speicher_chargeState = pd.DataFrame(index=self.calc.timeSeries)
        df_speicher_nettoFlow = pd.DataFrame(index=self.calc.timeSeries)

        list_of_speicher = [comp.label for comp in self.calc.comp_posts if comp.type == "cStorage"]

        for comp in list_of_speicher:
            df_speicher_chargeState[comp] = self.calc.results[comp]["charge_state"][:-1]  # without the last step
            df_speicher_nettoFlow[comp] = self.calc.results[comp]["nettoFlow"]

        if allocated:
            charge_state_sum = df_speicher_chargeState.sum(axis=1)
            netto_flow_sum = df_speicher_nettoFlow.sum(axis=1)*-1

            df = pd.concat([charge_state_sum, netto_flow_sum], axis=1)
            df.columns = ["Gesamtspeicherstand", "Nettospeicherflow"]
            df = resample_data(df, self.calc.years, resamply_by, rs_method)
        else:
            df = resample_data(df_speicher_chargeState, self.calc.years, resamply_by, rs_method)

        return df

    def get_fernwaerme_last_and_loss(self, resamply_by, rs_method):
        df_demand = self.calc.to_dataFrame("Waermelast", "in")
        df_loss = self.calc.to_dataFrame("Waermelast_Netzverluste", "in")
        df = pd.concat([df_demand, df_loss], axis=1)
        df_summed = resample_data(df, self.calc.years, resamply_by, rs_method)
        df_verluste_summed = (df_summed.iloc[:, 1] / df_summed.sum(axis=1) * 100).rename("Verlust[%]").round(2)

        return pd.concat([df_summed, df_verluste_summed], axis=1)

    def get_speicher_flows(self, resamply_by, rs_method, allocated):
        '''
        Parameters
        ----------
        resamply_by : string
            "h" for hourly resampling
            "d" for daily resampling
            "YE" for yearly resampling
        rs_method : string
            "mean" for mean value
            "sum" for sum value
            "max" for max value
            "min" for min value
        allocated : boolean
            True: sum of all storages in column one and netto flow of all Storages in column 2
            False: charge state of storage separately, no flow values

        Returns
        -------
        resampled DataFrame with total charge_state of all Storages
        '''
        df_speicher_nettoFlow = pd.DataFrame(index=self.calc.timeSeries)

        list_of_speicher = [comp.label for comp in self.calc.comp_posts if comp.type == "cStorage"]

        for comp in list_of_speicher:
            df_speicher_nettoFlow[comp] = self.calc.results[comp]["nettoFlow"]*-1

        if allocated:
            df = df_speicher_nettoFlow.sum(axis=1)
            df = resample_data(df, self.calc.years, resamply_by, rs_method)
            df.columns = ["Nettospeicherflow"]
        else:
            df = resample_data(df_speicher_nettoFlow, self.calc.years, resamply_by, rs_method)

        return df

    def merge_into_dispatch_structure(self, df:pd.DataFrame) -> pd.DataFrame:
        '''
        Brings a dataframe into a predefined structure for dispatch evaluation.
        Has space for 9 undefined columns
        '''
        # Step 1: Create an empty DataFrame with specific column names
        fixed_columns_1 = ['TAB', 'Geothermie', 'Abwärme', 'WP', 'WP_2', 'EHK', 'KWK_Gas', 'KWK_H2',
                           'Kessel_Gas', 'Kessel_H2', 'Speicher_S', 'Speicher_L', 'Kühler']  # First 11 fixed columns
        undefined_columns = ['U1', 'U2', 'U3', 'U4', 'U5', 'U6', 'U7', 'U8', 'U9']  # 8 undefined placeholders
        fixed_columns_2 = ['others', 'Wärmelast', 'Strompreis']  # Last 2 fixed columns

        # Combine all parts into the final column structure
        all_columns = fixed_columns_1 + undefined_columns + fixed_columns_2

        # Step 2: Create the target DataFrame with this structure, initially filled with None
        df_target = pd.DataFrame(columns=all_columns, index=df.index)

        # String formattin to prevent unintended behaviour
        df.columns = (df.columns
                      .str.replace('ae', 'ä')
                      .str.replace('oe', 'ö')
                      .str.replace('ue', 'ü')
                      .str.strip()
                      )
        df.columns = [col[0].upper() + col[1:] for col in df.columns]

        df.rename(columns={"Wärmelast_mit_Verlust": "Wärmelast"}, inplace=True)

        # Merge logic
        # Directly assign matched columns
        for col in df.columns.intersection(df_target.columns):
            df_target[col] = df[col]

        # Handle unmatched columns by placing them into the undefined placeholders
        unmatched_columns = df.columns.difference(df_target.columns)
        unmatched_columns = sorted(unmatched_columns, key=lambda x: x.lower())  # sorting alphabetically
        for i, col in enumerate(unmatched_columns):
            if i < len(undefined_columns):  # Ensure there's an available placeholder
                df_target[undefined_columns[i]] = df[col]
                df_target = df_target.rename(columns={undefined_columns[i]: col})
            else:
                df_target['others'] = df[[col for col in unmatched_columns[i:]]].sum(axis=1)

        # removing all values when all nan values
        nan_columns = df_target.columns[df_target.isnull().all()]
        rename_dict = {col: f"_{i}" for i, col in enumerate(nan_columns)}
        df_target = df_target.rename(columns=rename_dict)
        return df_target

#TODO: Adjust cExcelFcts() in a way, that other effects, other bus names and so on can be used.
# TODO: Evaluate, if even usefull
