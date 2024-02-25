import pandas as pd
import numpy as np
import os.path
from typing import Literal, List, Union
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference,LineChart
from openpyxl.utils.dataframe import dataframe_to_rows


def resample_data(data_frame: Union[pd.DataFrame, np.ndarray], target_years: List[int], resampling_by: Literal["YE", "d", "h"],
                   resampling_method: Literal["sum", "mean", "min","max"], initial_sampling_rate: str = "h") -> pd.DataFrame:
    '''
    Parameters
    ----------
    data_frame : Union[pd.DataFrame, np.ndarray]
        DataFrame or array containing data. Number of rows must match the initial sampling rate (safety check):
        8760 ("h") (default) or 365 ("d") per year.
    target_years : List[int]
        Target years for the new index of the DataFrame
    resampling_by : str
        "h" for hourly resampling
        "d" for daily resampling
        "YE" for yearly resampling
    resampling_method : str
        "mean" for mean value
        "sum" for sum value
        "max" for max value
        "min" for min value
    initial_sampling_rate : str
        "h" for hourly data (8760 values per year)
        "d" for daily data (365 values per year)

    Returns
    -------
    pd.DataFrame
    '''
    df = pd.DataFrame(data_frame)
    df.index = range(len(df))  # reset index

    if len(df)/8760 == len(target_years) and initial_sampling_rate == "h":
        length_per_year = 8760
    elif len(df)/365 == len(target_years) and initial_sampling_rate == "d":
        length_per_year = 365
    else:
        raise ValueError("length of dataframe and initial_sampling_rate must match: "
                         "8760 rows/year ('H') or 365 rows/year 'D'.")

    if not isinstance(target_years, list):
        target_years = [target_years]

    # create new TS for resampling, without the 29. February (filtering leap years)
    for i, year in enumerate(target_years):
        dt = pd.date_range(start=f'1/1/{year}', end=f'01/01/{year + 1}', freq=initial_sampling_rate)[:-1]
        dt = dt[~((dt.month == 2) & (dt.day == 29))]  # Remove leap year days
        df.loc[i * length_per_year:(i + 1) * length_per_year - 1, 'Timestamp'] = dt
    df = df.set_index('Timestamp')

    if resampling_method == "sum":
        df = df.resample(resampling_by).sum()
    elif resampling_method == "mean":
        df = df.resample(resampling_by).mean()
    elif resampling_method == "min":
        df = df.resample(resampling_by).min()
    elif resampling_method == "max":
        df = df.resample(resampling_by).max()
    else:
        raise ValueError("Invalid resampling method")

    # Drop all rows that aren't in the years specified in target_years
    lst = [row for row in df.index if row.year not in target_years]
    df = df.drop(index=lst)

    df = df.loc[~((df.index.month == 2) & (df.index.day == 29))]  # Remove leap year days again

    if resampling_by == "YE":
        df = df.set_index(df.index.year)  # setting the index to the plain year. No datetime anymore

    return df

def rs_in_two_steps(data_frame: Union[pd.DataFrame, np.ndarray], target_years: List[int], resampling_by: Literal["d", "YE"],
                    initial_sampling_rate: str = "h") -> pd.DataFrame:
    '''
    Parameters
    ----------
    data_frame : Union[pd.DataFrame, np.ndarray]
        DataFrame or array containing data. Number of rows must match the initial sampling rate (safety check):
        8760 ("h") (default) or 365 ("d") per year.
    target_years : List[int]
        Years for resampling
    resampling_by : str
        "d" for daily resampling
        "YE" for yearly resampling
    initial_sampling_rate : str
        "h" for hourly data
        "d" for daily data
    Returns
    -------
    pd.DataFrame
        Resampled DataFrame with new columns:
        ["Tagesmittel", "Minimum (Stunde)", "Maximum (Stunde)"]
        or new Columns:
        ["Jahresmittel", "Minimum (Tagesmittel)", "Maximum (Tagesmittel)"],
        depending on chosen "resampling_by"
    '''

    # Determine base resampling method and new columns based on resampling_by
    if resampling_by == "d":
        rs_method_base = "h"
        new_columns = ["Tagesmittel", "Minimum (Stunde)", "Maximum (Stunde)"]
    elif resampling_by == "YE":
        rs_method_base = "d"
        new_columns = ["Jahresmittel", "Minimum (Tagesmittel)", "Maximum (Tagesmittel)"]
    else:
        raise ValueError("Invalid value for resampling_by. Use 'D' for daily or 'Y' for yearly.")


    # Base resampling
    df_resampled_base = resample_data(data_frame, target_years, rs_method_base, "mean", initial_sampling_rate)

    # Resample for min, max, and mean
    min_y = resample_data(df_resampled_base, target_years, resampling_by, "min", rs_method_base)
    max_y = resample_data(df_resampled_base, target_years, resampling_by, "max", rs_method_base)
    mean_y = resample_data(df_resampled_base, target_years, resampling_by, "mean", rs_method_base)

    # Concatenate results
    df_result = pd.concat([mean_y, min_y, max_y], axis=1)
    df_result.columns = new_columns

    return df_result

def reorder_columns(df:pd.DataFrame, not_sorted_columns: List[str] = None):
    '''
    Order a DataFrame by a custom function, excluding specified columns from sorting, and setting them as the first columns.

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame.
    not_sorted_columns : List[str], optional
        Columns to exclude from sorting and set as the first columns, by default None.

    Returns
    -------
    pd.DataFrame
        DataFrame with the desired column order.
    '''
    if isinstance(df, pd.Series): df = df.to_frame().T
    sorted_columns = sorted(df.columns, key=lambda x: x.lower())
    sorted_df = df.reindex(columns=sorted_columns)

    # Select the remaining columns excluding the first two
    if not_sorted_columns is None:
        other_columns = [col for col in sorted_df.columns]
        # Create a new DataFrame with the desired column order
        new_order_df = sorted_df[other_columns]
    else:
        other_columns = [col for col in sorted_df.columns if col not in not_sorted_columns]

        # Create a new DataFrame with the desired column order
        new_order_df = pd.concat([sorted_df[not_sorted_columns], sorted_df[other_columns]], axis=1)

    return new_order_df

def df_to_excel_w_chart(df: pd.DataFrame, filepath: str, title: str, ylabel: str, xlabel: str, style:Literal["bar","line"]="bar"):
    """
    Write DataFrame to an Excel file with a stacked bar chart.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame containing the data to be written.
    filepath : str
        The path to the Excel file. If the file doesn't exist, a new one will be created.
    title : str
        The title of the sheet and chart.
    ylabel : str
        The label for the y-axis of the chart.
    xlabel : str
        The label for the x-axis of the chart.

    Returns
    -------
    None

    Notes
    -----
    This function writes the provided DataFrame to an Excel file and adds a stacked bar chart to a new sheet in the workbook.
    If the sheet with the given title already exists, it is removed before adding the new sheet.
    The stacked bar chart is created based on the DataFrame structure, with columns as categories and rows as data points.
    The chart is positioned at cell "D4" in the sheet.

    """
    try:
        wb = load_workbook(filepath)
    except FileNotFoundError:
        template_path = os.path.join( os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                         "resources", "Template_blanco.xlsx")
        wb = load_workbook(template_path)

    # Check if the sheet already exists
    if title in wb.sheetnames:
        sheet = wb[title]
        wb.remove(sheet)

    # Add the sheet to the workbook
    sheet = wb.create_sheet(title)

    # Remove the index and save it as a column
    df = df.reset_index()
    # Write the data starting from the second row
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # Create the Data and References
    data = Reference(sheet, min_col=2, min_row=1, max_col=df.shape[1], max_row=df.shape[0] + 1)
    labels = Reference(sheet, min_col=1, min_row=2, max_row=df.shape[0] + 1)

    # Create a stacked bar chart
    if style=="bar":
        chart = BarChart()
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        # Stacked bar plot
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.gapWidth = 0  # Adjust the gap between bars (e.g., set gapWidth to 0%)
    elif style=="line":
        chart = LineChart()
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        # Stacked bar plot
        chart.type = "line"

    # General Chart stuff
    chart.title = title
    chart.y_axis.title = ylabel
    chart.x_axis.title = xlabel
    chart.width = 30
    chart.height = 15

    # Add the chart to the sheet
    sheet.add_chart(chart, "D4")  # Adjust the position as needed

    # Save the workbook
    wb.save(filepath)

